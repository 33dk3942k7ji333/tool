from typing import cast

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

data = {
    "ID": range(1, 11),
    "產線 A": ["Y132", "N", "Y", np.nan, "N", "Y", np.nan, "N", "Y", "N"],
    "產線 B": [np.nan, "Y", "N", "N", "Y", np.nan, "Y", "N", "Y", "Y"],
    "產線 C": ["N", "N", np.nan, "Y", "Y", "N", "Y", np.nan, "N", "N"],
    "備註": ["正常", "測試", np.nan, "維修", "正常", np.nan, "正常", "測試", "正常", "維修"],
}
df = pd.DataFrame(data)

file_name = "large_styled_report.xlsx"
df.to_excel(file_name, index=False, engine="openpyxl")
wb = load_workbook(file_name)
ws = cast(Worksheet, wb.active)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, str):
            val = cell.value.upper()
            if "Y" in val:
                cell.style = "Bad"
            elif "N" in val:
                cell.style = "Good"
wb.save(file_name)
