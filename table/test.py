from typing import Any, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

Coordinate = Tuple[int, int]


def write_value(ws: Worksheet, coord: Coordinate, value: Any):
    ws.cell(row=coord[0], column=coord[1], value=value)


def merge_cells(ws: Worksheet, start_pos: Coordinate, end_pos: Coordinate):
    ws.merge_cells(start_row=start_pos[0], start_column=start_pos[1], end_row=end_pos[0], end_column=end_pos[1])


def set_center_alignment(ws: Worksheet, start_pos: Coordinate, end_pos: Coordinate):
    for row in ws.iter_rows(min_row=start_pos[0], max_row=end_pos[0], min_col=start_pos[1], max_col=end_pos[1]):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )


def set_bold_font(ws: Worksheet, start_pos: Coordinate, end_pos: Coordinate):
    for row in ws.iter_rows(min_row=start_pos[0], max_row=end_pos[0], min_col=start_pos[1], max_col=end_pos[1]):
        for cell in row:
            cell.font = Font(bold=True)


def set_outer_border(ws: Worksheet, start_pos: Coordinate, end_pos: Coordinate):
    thin = Side(border_style="thin")

    start_row, start_col = start_pos
    end_row, end_col = end_pos
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            current_border = cell.border

            left = thin if c == start_col else current_border.left
            right = thin if c == end_col else current_border.right
            top = thin if r == start_row else current_border.top
            bottom = thin if r == end_row else current_border.bottom

            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def call_format(
    ws: Worksheet,
    start: Coordinate,
    end: Coordinate,
    merge: bool = False,
    center: bool = False,
    bold: bool = False,
    border: bool = False,
):
    if merge:
        merge_cells(ws, start, end)
    if center:
        set_center_alignment(ws, start, end)
    if bold:
        set_bold_font(ws, start, end)
    if border:
        set_outer_border(ws, start, end)


def main():
    df1 = pd.DataFrame(
        {
            "TYPE": ["SOC"] * 5 + ["CPE"] * 5,
            "FAB": [f"FAB{i}" for i in range(1, 6)] + [f"FAB{i}" for i in range(1, 6)],
            "C": range(10, 20),
            "D": range(20, 30),
            "E": range(30, 40),
            "F": range(40, 50),
        }
    )

    df2 = pd.DataFrame(
        {
            "TYPE": ["SOC"] * 5 + ["CPE"] * 5,
            "FAB": [f"FAB{i}" for i in range(1, 6)] + [f"FAB{i}" for i in range(1, 6)],
            "C": range(50, 60),
            "D": range(60, 70),
            "E": range(70, 80),
            "F": range(80, 90),
        }
    )

    file_name: str = "styled_report.xlsx"
    sheet_name: str = "Report"

    sr_idx = 0
    cr_idx = 12
    lst_time = ["C", "D", "E", "F"]
    lst_fab = [1, 2, 3, 4, 5]

    num_time = len(lst_time)
    num_fab = len(lst_fab)
    num_w = 2
    num_m = 2

    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df1.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0, startcol=sr_idx)
        df2.to_excel(writer, index=False, sheet_name=sheet_name, startrow=0, startcol=cr_idx)

        ws: Worksheet = writer.sheets[sheet_name]

        # write table title
        cell_sr = (1, sr_idx + 1)
        cell_cr = (1, cr_idx + 1)
        write_value(ws, cell_sr, "SR")
        write_value(ws, cell_cr, "CR")

        # merge type
        p1_soc, p2_soc = (2 + 0 * num_fab, sr_idx + 1), (2 + 1 * num_fab - 1, sr_idx + 1)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, sr_idx + 1), (2 + 2 * num_fab - 1, sr_idx + 1)
        call_format(ws, p1_soc, p2_soc, merge=True, center=True, bold=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, merge=True, center=True, bold=True, border=True)

        p1_soc, p2_soc = (2 + 0 * num_fab, cr_idx + 1), (2 + 1 * num_fab - 1, cr_idx + 1)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, cr_idx + 1), (2 + 2 * num_fab - 1, cr_idx + 1)
        call_format(ws, p1_soc, p2_soc, merge=True, center=True, bold=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, merge=True, center=True, bold=True, border=True)

        # merge fab
        p1_soc, p2_soc = (2 + 0 * num_fab, sr_idx + 2), (2 + 1 * num_fab - 1, sr_idx + 2)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, sr_idx + 2), (2 + 2 * num_fab - 1, sr_idx + 2)
        call_format(ws, p1_soc, p2_soc, center=True, bold=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, center=True, bold=True, border=True)

        p1_soc, p2_soc = (2 + 0 * num_fab, cr_idx + 2), (2 + 1 * num_fab - 1, cr_idx + 2)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, cr_idx + 2), (2 + 2 * num_fab - 1, cr_idx + 2)
        call_format(ws, p1_soc, p2_soc, center=True, bold=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, center=True, bold=True, border=True)

        # merge week
        p1_soc, p2_soc = (2 + 0 * num_fab, sr_idx + 3), (2 + 1 * num_fab - 1, sr_idx + 3 + num_w - 1)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, sr_idx + 3), (2 + 2 * num_fab - 1, sr_idx + 3 + num_w - 1)
        call_format(ws, p1_soc, p2_soc, center=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, center=True, border=True)

        p1_soc, p2_soc = (2 + 0 * num_fab, cr_idx + 3), (2 + 1 * num_fab - 1, cr_idx + 3 + num_w - 1)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, cr_idx + 3), (2 + 2 * num_fab - 1, cr_idx + 3 + num_w - 1)
        call_format(ws, p1_soc, p2_soc, center=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, center=True, border=True)

        # merge month
        p1_soc, p2_soc = (2 + 0 * num_fab, sr_idx + 3 + num_w), (2 + 1 * num_fab - 1, sr_idx + 3 + num_time - 1)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, sr_idx + 3 + num_w), (2 + 2 * num_fab - 1, sr_idx + 3 + num_time - 1)
        call_format(ws, p1_soc, p2_soc, center=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, center=True, border=True)

        p1_soc, p2_soc = (2 + 0 * num_fab, cr_idx + 3 + num_w), (2 + 1 * num_fab - 1, cr_idx + 3 + num_time - 1)
        p1_cpe, p2_cpe = (2 + 1 * num_fab, cr_idx + 3 + num_w), (2 + 2 * num_fab - 1, cr_idx + 3 + num_time - 1)
        call_format(ws, p1_soc, p2_soc, center=True, border=True)
        call_format(ws, p1_cpe, p2_cpe, center=True, border=True)


if __name__ == "__main__":
    main()
